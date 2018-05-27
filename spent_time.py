import datetime
import os
from openpyxl import Workbook
from redmine import Redmine
from sys import exit
from openpyxl.styles import PatternFill
from auth import *
from requests.exceptions import ConnectionError

wb = Workbook()
ws1 = wb.active
colorFill = PatternFill(start_color='0073C2FB', end_color='0073C2FB', fill_type='solid')
start_date = datetime.datetime.now() - datetime.timedelta(7)


total_spent_time = []
xlsx_data = []



def create_report():
    try:
        redmine = Redmine(redmine_url, key = redmine_api_key)
    except:
        return False
    user = redmine.user.get(user_id)
    spent_time = user.time_entries
    user_for_xlsx = user.lastname + ' ' + user.firstname
    dest_filename = os.path.join("reports", start_date.strftime("%d.%m.%Y") + "-" + datetime.datetime.now().strftime("%d.%m.%Y") + "_report_" + user.login + ".xlsx")
    for entry in spent_time:
        delta = (datetime.datetime.now() - entry.created_on).days
        if delta <= 7:
            iss_subj = redmine.issue.get(entry.issue.id).subject
            iss_project = redmine.issue.get(entry.issue.id).project.name
            task = str(entry.issue.id) + ' ' + iss_subj
            total_spent_time.append(entry.hours)
            xlsx_data.append([entry.created_on.strftime("%d-%m-%Y"), user_for_xlsx, 'Tech', iss_project, task, float(entry.hours)])



    ws1['A1'] = "Date"
    ws1['B1'] = "Name"
    ws1['C1'] = "Department"
    ws1['D1'] = "Project"
    ws1['E1'] = "Task"
    ws1['F1'] = "Hours"

    ws1['A1'].fill = colorFill
    ws1['B1'].fill = colorFill
    ws1['C1'].fill = colorFill
    ws1['D1'].fill = colorFill
    ws1['E1'] .fill= colorFill
    ws1['F1'] .fill= colorFill

    for row in xlsx_data:
        ws1.append(row)

    for col in ws1.columns:
         max_length = 0
         column = col[0].column 
         for cell in col:
             try: 
                 if len(str(cell.value)) > max_length:
                     max_length = len(cell.value)
             except:
                 pass
         adjusted_width = (max_length + 2) * 2
         ws1.column_dimensions[column].width = adjusted_width

    for i in range(1,100):
        try:
            cell = 'E%s' % i
            len(ws1[cell].value)
        except TypeError:
            cell_F = 'F%s' % i
            ws1[cell] = 'Total'
            ws1[cell_F] = float(sum(total_spent_time))
            wb.save(filename = dest_filename)
            return dest_filename


